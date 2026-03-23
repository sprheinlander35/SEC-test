#!/usr/bin/env python3
"""
SEC EDGAR 10-K Financial Statement Extractor
=============================================
Fetches XBRL-tagged annual 10-K data from SEC EDGAR and exports
three financial statements (Income Statement, Balance Sheet, Cash Flow)
to a formatted Excel workbook.

Usage:
    python sec_extractor.py AAPL
    python sec_extractor.py MSFT --years 15
    python sec_extractor.py TSLA --output ./reports
    python sec_extractor.py           (will prompt for ticker)
"""

import sys
import time
import argparse
import requests
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── SEC EDGAR API configuration ─────────────────────────────────────────────
# SEC requires a descriptive User-Agent string (not a browser UA)
HEADERS = {
    "User-Agent": "SEC-10K-Extractor research@example.com",
    "Accept-Encoding": "gzip, deflate",
}
EDGAR_BASE = "https://www.sec.gov"
DATA_BASE  = "https://data.sec.gov"


# ── Step 1: CIK lookup ───────────────────────────────────────────────────────

def lookup_cik(ticker: str) -> tuple:
    """Return (zero-padded CIK string, company name) for a ticker."""
    url = f"{EDGAR_BASE}/files/company_tickers.json"
    r = requests.get(url, headers={"User-Agent": HEADERS["User-Agent"]}, timeout=30)
    r.raise_for_status()

    ticker_upper = ticker.upper()
    for entry in r.json().values():
        if entry["ticker"].upper() == ticker_upper:
            cik = str(entry["cik_str"]).zfill(10)
            return cik, entry["title"]

    raise ValueError(
        f"Ticker '{ticker}' not found in SEC EDGAR. "
        f"Verify the symbol at https://www.sec.gov/cgi-bin/browse-edgar"
    )


# ── Step 2: Fetch XBRL company facts ────────────────────────────────────────

def fetch_company_facts(cik: str) -> dict:
    """Download all XBRL facts for a company (can be 10–100 MB)."""
    url = f"{DATA_BASE}/api/xbrl/companyfacts/CIK{cik}.json"
    r = requests.get(url, headers=HEADERS, timeout=120)
    r.raise_for_status()
    return r.json()


# ── Step 3: Extract annual series ────────────────────────────────────────────

def extract_series(facts: dict, namespace: str, concept: str) -> dict:
    """
    Extract {fiscal_year (int): value} for a XBRL concept.
    Keeps only annual 10-K data (fp='FY', form='10-K').
    Deduplicates by keeping the most recently filed revision per FY.
    Extracts year from 'end' date field when 'fy' is absent (older filings).
    """
    try:
        units = facts["facts"][namespace][concept]["units"]
        # Prefer USD; fall back to shares or any other unit
        items = (
            units.get("USD")
            or units.get("shares")
            or next(iter(units.values()))
        )

        best: dict = {}
        for item in items:
            if item.get("form") not in ("10-K", "10-K/A"):
                continue
            if item.get("fp") != "FY":
                continue

            # Use 'fy' field when present; otherwise parse year from end date
            fy = item.get("fy")
            if not fy and item.get("end"):
                try:
                    fy = int(item["end"][:4])
                except (ValueError, TypeError):
                    continue
            if not fy:
                continue

            filed = item.get("filed", "")
            if fy not in best or filed > best[fy].get("filed", ""):
                best[fy] = item

        return {fy: d["val"] for fy, d in sorted(best.items())}

    except (KeyError, TypeError, StopIteration):
        return {}


def first_with_data(facts: dict, candidates: list) -> dict:
    """Try XBRL concept alternatives in order; return the first that has data."""
    for namespace, concept in candidates:
        data = extract_series(facts, namespace, concept)
        if data:
            return data
    return {}


# ── XBRL concept definitions ─────────────────────────────────────────────────
# Each entry: (row label, [(namespace, concept), ...])
# Multiple alternatives are tried in order; the first with data wins.

INCOME_STMT = [
    ("Revenue", [
        ("us-gaap", "Revenues"),
        ("us-gaap", "RevenueFromContractWithCustomerExcludingAssessedTax"),
        ("us-gaap", "RevenueFromContractWithCustomerIncludingAssessedTax"),
        ("us-gaap", "SalesRevenueNet"),
        ("us-gaap", "SalesRevenueGoodsNet"),
        ("us-gaap", "RevenuesNetOfInterestExpense"),
        ("us-gaap", "BankingAndThriftDisclosureTextBlock"),  # banks
    ]),
    ("Cost of Revenue", [
        ("us-gaap", "CostOfRevenue"),
        ("us-gaap", "CostOfGoodsAndServicesSold"),
        ("us-gaap", "CostOfGoodsSold"),
        ("us-gaap", "CostOfServices"),
    ]),
    ("Gross Profit", [
        ("us-gaap", "GrossProfit"),
    ]),
    ("R&D Expenses", [
        ("us-gaap", "ResearchAndDevelopmentExpense"),
        ("us-gaap", "ResearchAndDevelopmentExpenseExcludingAcquiredInProcessCost"),
    ]),
    ("SG&A Expenses", [
        ("us-gaap", "SellingGeneralAndAdministrativeExpense"),
        ("us-gaap", "GeneralAndAdministrativeExpense"),
    ]),
    ("Operating Income / (Loss)", [
        ("us-gaap", "OperatingIncomeLoss"),
    ]),
    ("Interest Expense", [
        ("us-gaap", "InterestExpense"),
        ("us-gaap", "InterestExpenseDebt"),
        ("us-gaap", "InterestAndDebtExpense"),
    ]),
    ("Pre-Tax Income / (Loss)", [
        ("us-gaap", "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest"),
        ("us-gaap", "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments"),
    ]),
    ("Income Tax Expense / (Benefit)", [
        ("us-gaap", "IncomeTaxExpenseBenefit"),
    ]),
    ("Net Income / (Loss)", [
        ("us-gaap", "NetIncomeLoss"),
        ("us-gaap", "NetIncomeLossAvailableToCommonStockholdersBasic"),
        ("us-gaap", "ProfitLoss"),
    ]),
    ("EPS — Basic ($ per share)", [
        ("us-gaap", "EarningsPerShareBasic"),
    ]),
    ("EPS — Diluted ($ per share)", [
        ("us-gaap", "EarningsPerShareDiluted"),
    ]),
    ("Shares — Basic (millions)", [
        ("us-gaap", "WeightedAverageNumberOfSharesOutstandingBasic"),
    ]),
    ("Shares — Diluted (millions)", [
        ("us-gaap", "WeightedAverageNumberOfDilutedSharesOutstanding"),
    ]),
]

BALANCE_SHEET = [
    ("Cash & Cash Equivalents", [
        ("us-gaap", "CashAndCashEquivalentsAtCarryingValue"),
        ("us-gaap", "Cash"),
        ("us-gaap", "CashAndDueFromBanks"),
    ]),
    ("Short-term Investments", [
        ("us-gaap", "ShortTermInvestments"),
        ("us-gaap", "MarketableSecuritiesCurrent"),
        ("us-gaap", "AvailableForSaleSecuritiesCurrent"),
    ]),
    ("Accounts Receivable, Net", [
        ("us-gaap", "AccountsReceivableNetCurrent"),
        ("us-gaap", "ReceivablesNetCurrent"),
    ]),
    ("Inventories", [
        ("us-gaap", "InventoryNet"),
        ("us-gaap", "InventoryFinishedGoodsAndWorkInProcess"),
        ("us-gaap", "InventoryFinishedGoods"),
    ]),
    ("Other Current Assets", [
        ("us-gaap", "OtherAssetsCurrent"),
        ("us-gaap", "PrepaidExpenseAndOtherAssetsCurrent"),
    ]),
    ("Total Current Assets", [
        ("us-gaap", "AssetsCurrent"),
    ]),
    ("PP&E, Net", [
        ("us-gaap", "PropertyPlantAndEquipmentNet"),
    ]),
    ("Goodwill", [
        ("us-gaap", "Goodwill"),
    ]),
    ("Intangible Assets, Net", [
        ("us-gaap", "IntangibleAssetsNetExcludingGoodwill"),
        ("us-gaap", "FiniteLivedIntangibleAssetsNet"),
    ]),
    ("Long-term Investments", [
        ("us-gaap", "LongTermInvestments"),
        ("us-gaap", "MarketableSecuritiesNoncurrent"),
        ("us-gaap", "AvailableForSaleSecuritiesNoncurrent"),
    ]),
    ("Other Non-current Assets", [
        ("us-gaap", "OtherAssetsNoncurrent"),
    ]),
    ("Total Assets", [
        ("us-gaap", "Assets"),
    ]),
    ("Accounts Payable", [
        ("us-gaap", "AccountsPayableCurrent"),
    ]),
    ("Short-term Debt", [
        ("us-gaap", "ShortTermBorrowings"),
        ("us-gaap", "DebtCurrent"),
        ("us-gaap", "CommercialPaper"),
        ("us-gaap", "NotesPayableCurrent"),
    ]),
    ("Accrued Liabilities", [
        ("us-gaap", "AccruedLiabilitiesCurrent"),
        ("us-gaap", "OtherLiabilitiesCurrent"),
    ]),
    ("Deferred Revenue (Current)", [
        ("us-gaap", "DeferredRevenueCurrent"),
        ("us-gaap", "ContractWithCustomerLiabilityCurrent"),
    ]),
    ("Total Current Liabilities", [
        ("us-gaap", "LiabilitiesCurrent"),
    ]),
    ("Long-term Debt", [
        ("us-gaap", "LongTermDebtNoncurrent"),
        ("us-gaap", "LongTermDebt"),
        ("us-gaap", "LongTermNotesPayable"),
    ]),
    ("Deferred Tax Liabilities", [
        ("us-gaap", "DeferredIncomeTaxLiabilitiesNet"),
        ("us-gaap", "DeferredTaxLiabilitiesNoncurrent"),
    ]),
    ("Other Non-current Liabilities", [
        ("us-gaap", "OtherLiabilitiesNoncurrent"),
    ]),
    ("Total Liabilities", [
        ("us-gaap", "Liabilities"),
    ]),
    ("Common Stock & APIC", [
        ("us-gaap", "CommonStocksIncludingAdditionalPaidInCapital"),
        ("us-gaap", "AdditionalPaidInCapital"),
        ("us-gaap", "AdditionalPaidInCapitalCommonStock"),
    ]),
    ("Retained Earnings / (Deficit)", [
        ("us-gaap", "RetainedEarningsAccumulatedDeficit"),
    ]),
    ("Accumulated Other Comprehensive Income", [
        ("us-gaap", "AccumulatedOtherComprehensiveIncomeLossNetOfTax"),
    ]),
    ("Total Stockholders' Equity", [
        ("us-gaap", "StockholdersEquity"),
        ("us-gaap", "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"),
    ]),
    ("Total Liabilities & Equity", [
        ("us-gaap", "LiabilitiesAndStockholdersEquity"),
    ]),
]

CASH_FLOW = [
    ("Net Income / (Loss)", [
        ("us-gaap", "NetIncomeLoss"),
        ("us-gaap", "ProfitLoss"),
    ]),
    ("Depreciation & Amortization", [
        ("us-gaap", "DepreciationDepletionAndAmortization"),
        ("us-gaap", "DepreciationAndAmortization"),
        ("us-gaap", "Depreciation"),
    ]),
    ("Stock-based Compensation", [
        ("us-gaap", "ShareBasedCompensation"),
        ("us-gaap", "AllocatedShareBasedCompensationExpense"),
    ]),
    ("Deferred Income Taxes", [
        ("us-gaap", "DeferredIncomeTaxExpenseBenefit"),
        ("us-gaap", "DeferredIncomeTaxesAndTaxCredits"),
    ]),
    ("Changes in Working Capital", [
        ("us-gaap", "IncreaseDecreaseInOperatingCapital"),
        ("us-gaap", "IncreaseDecreaseInOperatingLiabilities"),
    ]),
    ("Operating Cash Flow", [
        ("us-gaap", "NetCashProvidedByUsedInOperatingActivities"),
        ("us-gaap", "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations"),
    ]),
    ("Capital Expenditures", [
        ("us-gaap", "PaymentsToAcquirePropertyPlantAndEquipment"),
        ("us-gaap", "PaymentsForCapitalImprovements"),
        ("us-gaap", "AcquisitionsNetOfCashAcquiredAndPurchasesOfBusinesses"),
    ]),
    # Free Cash Flow is calculated and inserted here
    ("Acquisitions, Net of Cash", [
        ("us-gaap", "PaymentsToAcquireBusinessesNetOfCashAcquired"),
        ("us-gaap", "PaymentsToAcquireBusinessesGross"),
    ]),
    ("Purchases of Investments", [
        ("us-gaap", "PaymentsToAcquireAvailableForSaleSecurities"),
        ("us-gaap", "PaymentsToAcquireMarketableSecurities"),
        ("us-gaap", "PaymentsToAcquireInvestments"),
    ]),
    ("Proceeds from Sale of Investments", [
        ("us-gaap", "ProceedsFromSaleAndMaturityOfMarketableSecurities"),
        ("us-gaap", "ProceedsFromSaleAndMaturityOfAvailableForSaleSecurities"),
        ("us-gaap", "ProceedsFromSaleMaturityAndCollectionsOfInvestments"),
    ]),
    ("Investing Cash Flow", [
        ("us-gaap", "NetCashProvidedByUsedInInvestingActivities"),
        ("us-gaap", "NetCashProvidedByUsedInInvestingActivitiesContinuingOperations"),
    ]),
    ("Debt Proceeds", [
        ("us-gaap", "ProceedsFromIssuanceOfLongTermDebt"),
        ("us-gaap", "ProceedsFromIssuanceOfDebt"),
        ("us-gaap", "ProceedsFromDebtNetOfIssuanceCosts"),
    ]),
    ("Debt Repayments", [
        ("us-gaap", "RepaymentsOfLongTermDebt"),
        ("us-gaap", "RepaymentsOfDebt"),
    ]),
    ("Dividends Paid", [
        ("us-gaap", "PaymentsOfDividends"),
        ("us-gaap", "PaymentsOfDividendsCommonStock"),
    ]),
    ("Share Repurchases", [
        ("us-gaap", "PaymentsForRepurchaseOfCommonStock"),
        ("us-gaap", "PaymentsForRepurchaseOfEquity"),
    ]),
    ("Financing Cash Flow", [
        ("us-gaap", "NetCashProvidedByUsedInFinancingActivities"),
        ("us-gaap", "NetCashProvidedByUsedInFinancingActivitiesContinuingOperations"),
    ]),
    ("Net Change in Cash", [
        ("us-gaap", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect"),
        ("us-gaap", "CashAndCashEquivalentsPeriodIncreaseDecrease"),
        ("us-gaap", "NetCashProvidedByUsedInContinuingOperations"),
    ]),
]

# Rows that are NOT in USD (no divide-by-1000 scaling)
EPS_ROWS   = {"EPS — Basic ($ per share)", "EPS — Diluted ($ per share)"}
SHARE_ROWS = {"Shares — Basic (millions)", "Shares — Diluted (millions)"}


# ── Step 4: Build DataFrames ──────────────────────────────────────────────────

def get_fiscal_years(facts: dict, max_years: int = 10) -> list:
    """Detect available fiscal years by probing key concepts."""
    probe = [
        ("us-gaap", "Assets"),
        ("us-gaap", "NetIncomeLoss"),
        ("us-gaap", "NetCashProvidedByUsedInOperatingActivities"),
        ("us-gaap", "Revenues"),
        ("us-gaap", "RevenueFromContractWithCustomerExcludingAssessedTax"),
    ]
    years: set = set()
    for ns, c in probe:
        years.update(extract_series(facts, ns, c).keys())

    current_year = datetime.now().year
    available = sorted(y for y in years if isinstance(y, int) and y <= current_year)
    return available[-max_years:]


def build_df(facts: dict, concepts: list, years: list) -> pd.DataFrame:
    """
    Build a financial-statement DataFrame.
    Rows = line items, Columns = fiscal years (as strings).
    Values are scaled: USD → thousands; shares → millions; EPS unchanged.
    """
    rows = {}
    for label, candidates in concepts:
        data = first_with_data(facts, candidates) if candidates else {}
        row_values = {y: data.get(y) for y in years}

        # Scale
        if label in EPS_ROWS:
            pass  # already per-share, no scaling
        elif label in SHARE_ROWS:
            row_values = {
                y: (round(v / 1e6, 2) if v is not None else None)
                for y, v in row_values.items()
            }
        else:
            row_values = {
                y: (round(v / 1e3, 0) if v is not None else None)
                for y, v in row_values.items()
            }

        rows[label] = row_values

    df = pd.DataFrame(rows, index=years).T
    df.columns = [str(y) for y in df.columns]
    return df


def add_free_cash_flow(cf_df: pd.DataFrame) -> pd.DataFrame:
    """Calculate Free Cash Flow = Operating CF + CapEx (CapEx is typically negative)."""
    if "Operating Cash Flow" not in cf_df.index or "Capital Expenditures" not in cf_df.index:
        return cf_df

    ocf   = cf_df.loc["Operating Cash Flow"]
    capex = cf_df.loc["Capital Expenditures"]

    fcf = {}
    for col in cf_df.columns:
        o = ocf[col]
        c = capex[col]
        if o is not None and c is not None and pd.notna(o) and pd.notna(c):
            # CapEx is reported as a negative outflow; FCF = OCF + CapEx
            # If somehow it's stored as positive, subtract it
            fcf[col] = round(o + c if c <= 0 else o - c, 0)
        else:
            fcf[col] = None

    capex_pos = list(cf_df.index).index("Capital Expenditures")
    fcf_row = pd.DataFrame([fcf], index=["Free Cash Flow"])
    return pd.concat([cf_df.iloc[: capex_pos + 1], fcf_row, cf_df.iloc[capex_pos + 1 :]])


# ── Step 5: Excel export & formatting ────────────────────────────────────────

# Color palette
C_DARK_BLUE  = "1F4E79"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "D6E4F0"
C_ALT_ROW    = "EBF3FB"
C_WHITE      = "FFFFFF"
C_DARK_GRAY  = "404040"

# Row labels that receive bold + stronger fill treatment
BOLD_ROWS = {
    "Revenue", "Gross Profit", "Operating Income / (Loss)", "Net Income / (Loss)",
    "Total Current Assets", "Total Assets",
    "Total Current Liabilities", "Total Liabilities",
    "Total Stockholders' Equity", "Total Liabilities & Equity",
    "Operating Cash Flow", "Free Cash Flow",
    "Investing Cash Flow", "Financing Cash Flow",
}


def style_worksheet(ws, company_name: str, ticker: str, sheet_title: str):
    """Apply professional blue-themed formatting to a worksheet."""
    n_cols = ws.max_column

    # ── Insert 3 title rows above the pandas header ──
    ws.insert_rows(1, 3)

    title = f"{company_name}  ({ticker.upper()})  —  {sheet_title}"
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True, size=13, color=C_DARK_BLUE)

    note = "Values in USD Thousands unless noted  |  EPS in $/share  |  Shares in Millions"
    ws.cell(2, 1).value = note
    ws.cell(2, 1).font = Font(italic=True, size=9, color="666666")

    src = f"Source: SEC EDGAR XBRL Data   |   Generated: {datetime.now():%Y-%m-%d}"
    ws.cell(3, 1).value = src
    ws.cell(3, 1).font = Font(italic=True, size=8, color="AAAAAA")

    # Merge title cells across all columns
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)

    # ── Column header row (now row 4 after insert) ──
    hdr_fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    hdr_font = Font(bold=True, color=C_WHITE, size=10)
    for col in range(1, n_cols + 1):
        cell = ws.cell(4, col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(
            horizontal="left" if col == 1 else "center",
            vertical="center",
        )

    # ── Data rows (row 5 onward) ──
    alt_fill  = PatternFill("solid", fgColor=C_ALT_ROW)
    bold_fill = PatternFill("solid", fgColor=C_LIGHT_BLUE)

    data_idx = 0
    for row in range(5, ws.max_row + 1):
        label = ws.cell(row, 1).value
        is_bold = label in BOLD_ROWS
        is_alt  = data_idx % 2 == 1
        data_idx += 1

        row_fill = bold_fill if is_bold else (alt_fill if is_alt else None)

        for col in range(1, n_cols + 1):
            cell = ws.cell(row, col)

            if row_fill:
                cell.fill = row_fill

            if col == 1:
                cell.font      = Font(bold=is_bold, size=9, color=C_DARK_GRAY)
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           indent=0 if is_bold else 1)
            else:
                cell.font      = Font(bold=is_bold, size=9)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                v = cell.value
                if v is not None and isinstance(v, (int, float)) and pd.notna(v):
                    # EPS rows: show 2 decimal places; everything else: integer thousands
                    cell.number_format = (
                        "#,##0.00" if abs(v) < 1000 else "#,##0"
                    )

    # ── Column widths ──
    ws.column_dimensions["A"].width = 36
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # ── Freeze panes (header + label column) ──
    ws.freeze_panes = ws.cell(5, 2)

    # ── Row heights ──
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 26
    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 15


def export_to_excel(
    income_df: pd.DataFrame,
    balance_df: pd.DataFrame,
    cf_df: pd.DataFrame,
    company_name: str,
    ticker: str,
    output_path: str,
):
    """Write DataFrames to Excel then apply formatting."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        income_df.to_excel(writer, sheet_name="Income Statement",
                           index_label="Metric")
        balance_df.to_excel(writer, sheet_name="Balance Sheet",
                            index_label="Metric")
        cf_df.to_excel(writer, sheet_name="Cash Flow Statement",
                       index_label="Metric")

    wb = load_workbook(output_path)
    for sheet_name, title in [
        ("Income Statement",    "Income Statement"),
        ("Balance Sheet",       "Balance Sheet"),
        ("Cash Flow Statement", "Cash Flow Statement"),
    ]:
        if sheet_name in wb.sheetnames:
            style_worksheet(wb[sheet_name], company_name, ticker, title)

    wb.save(output_path)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="SEC EDGAR 10-K Financial Statement Extractor",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sec_extractor.py AAPL
  python sec_extractor.py MSFT --years 15
  python sec_extractor.py GOOGL --output ./reports
        """,
    )
    parser.add_argument("ticker", nargs="?",
                        help="Stock ticker symbol (e.g., AAPL, MSFT, TSLA)")
    parser.add_argument("--years", type=int, default=10,
                        help="Years of history to retrieve (default: 10)")
    parser.add_argument("--output", type=str, default=".",
                        help="Output directory (default: current directory)")
    args = parser.parse_args()

    ticker = (
        args.ticker or input("\nEnter stock ticker (e.g., AAPL, MSFT, TSLA): ")
    ).strip().upper()

    if not ticker:
        print("Error: no ticker provided.")
        sys.exit(1)

    print(f"\n{'='*65}")
    print("  SEC EDGAR 10-K Financial Statement Extractor")
    print(f"{'='*65}")

    # ── 1. CIK lookup ──────────────────────────────────────────────────
    print(f"\n[1/5] Looking up '{ticker}' in SEC EDGAR...")
    try:
        cik, company_name = lookup_cik(ticker)
    except ValueError as e:
        print(f"\n  ERROR: {e}")
        sys.exit(1)
    except requests.RequestException as e:
        print(f"\n  Network error during CIK lookup: {e}")
        sys.exit(1)

    print(f"      Company : {company_name}")
    print(f"      CIK     : {cik}")

    # ── 2. Download XBRL facts ─────────────────────────────────────────
    print(f"\n[2/5] Downloading XBRL data from SEC EDGAR (may take a moment)...")
    time.sleep(0.3)  # polite delay per SEC rate-limit guidelines
    try:
        facts = fetch_company_facts(cik)
    except requests.RequestException as e:
        print(f"\n  Error fetching XBRL data: {e}")
        sys.exit(1)

    concept_count = len(facts.get("facts", {}).get("us-gaap", {}))
    print(f"      Retrieved {concept_count:,} XBRL concepts")

    # ── 3. Identify fiscal years ───────────────────────────────────────
    print(f"\n[3/5] Identifying available fiscal years...")
    years = get_fiscal_years(facts, max_years=args.years)
    if not years:
        print("  ERROR: No annual 10-K XBRL data found for this company.")
        sys.exit(1)
    print(f"      Available : FY{years[0]} — FY{years[-1]}  ({len(years)} years)")

    # ── 4. Build financial statements ──────────────────────────────────
    print(f"\n[4/5] Extracting financial statements...")
    income_df  = build_df(facts, INCOME_STMT,   years)
    balance_df = build_df(facts, BALANCE_SHEET,  years)
    cf_df      = build_df(facts, CASH_FLOW,      years)
    cf_df      = add_free_cash_flow(cf_df)

    print(f"      Income Statement    : {len(income_df)} line items")
    print(f"      Balance Sheet       : {len(balance_df)} line items")
    print(f"      Cash Flow Statement : {len(cf_df)} line items")

    # ── 5. Export to Excel ─────────────────────────────────────────────
    print(f"\n[5/5] Exporting to Excel...")
    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(out_dir / f"{ticker}_10K_Financials.xlsx")

    export_to_excel(income_df, balance_df, cf_df, company_name, ticker, output_path)

    print(f"\n{'='*65}")
    print(f"  SUCCESS!")
    print(f"  File    : {output_path}")
    print(f"  Company : {company_name}")
    print(f"  Period  : FY{years[0]} — FY{years[-1]}")
    print(f"{'='*65}\n")


if __name__ == "__main__":
    main()
